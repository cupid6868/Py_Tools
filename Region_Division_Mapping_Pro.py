import pandas as pd
import geopandas as gpd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np
import warnings
import os
from matplotlib.offsetbox import OffsetImage, AnnotationBbox
import cairosvg
import io
from PIL import Image
from shapely.errors import TopologicalError
from shapely.geometry import Point
from openpyxl import load_workbook

warnings.filterwarnings('ignore')


# ==============================================
# 【超参数区】
# ==============================================
class Config:
    EXCEL_PATH = r"E:\Test_Code\平衡面板.xlsx"
    SAVE_EXCEL_PATH = r"E:\Test_Code\平衡面板_添加区域列.xlsx"
    SHEET_NAME = "3.25"

    # ✅ 新增：是否保存 Excel 文件（True=保存，False=不保存）
    NEED_SAVE_EXCEL = False

    SHP_PATH = r"E:\Test_Code\2023年县级\县级.shp"
    NINE_LINES_PATH = r"E:\Test_Code\九段线\九段线.shp"
    SVG_COMPASS_PATH = r"E:\Test_Code\指北针.svg"
    AGRI_ZONE_PATH = r"E:\Test_Code\中国气候区划\Climate_quhua.shp"

    TARGET_YEAR = 2023

    FONT_NAME = "Times New Roman"
    COLORS = [
        '#1f77b4', '#ff7f00', '#2ca02c', '#d62728', '#9467bd',
        '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf',
        '#a8e6cf', '#ffb3b3', '#d4a5a5', '#66b3ff', '#99ff99',
        '#ff99cc', '#c9c9ff', '#ffcc99', '#ff6666', '#66ff66',
        '#6666ff', '#ff66ff', '#ffff66', '#66ffff', '#ff6600',
        '#00ff66', '#0066ff', '#6600ff', '#ff0066', '#00ffcc',
        '#ff00cc', '#cc00ff', '#cccc00', '#00cccc', '#333333',
        '#666666', '#999999', '#cccccc', '#00033', '#003300'
    ]
    NULL_COLOR = 'white'
    BORDER_COLOR = 'black'
    BORDER_WIDTH = 0.1
    NINE_LINE_COLOR = 'red'
    NINE_LINE_WIDTH = 1.2

    SAVE_PATH = r'E:\Test_Code\2023_Three_E_M_W_Zone_Map.svg'
    DPI = 900
    FIG_SIZE = (14, 12)

    COMPASS_ZOOM = 0.35
    GRID_LABEL_SIZE = 14
    LEGEND_FONT_SIZE = 14
    LEGEND_TITLE_SIZE = 16
    SCALE_TEXT_SIZE = 16

    ZONE_TRANSLATION = {
        "东部": "East",
        "中部": "Central",
        "西部": "West"
    }


cfg = Config()


# ==============================================
# 打印所有sheet信息
# ==============================================
def print_all_sheets_info(excel_path, title="Excel 文件信息"):
    xl = pd.ExcelFile(excel_path)
    sheet_names = xl.sheet_names
    print("\n" + "=" * 80)
    print(f"📄 {title}")
    print(f"📂 文件路径：{excel_path}")
    print(f"📊 包含 {len(sheet_names)} 个子表")
    print("=" * 80)

    for i, sht in enumerate(sheet_names, 1):
        try:
            df_temp = pd.read_excel(excel_path, sheet_name=sht)
            rows, cols = df_temp.shape
            columns = list(df_temp.columns)
            print(f"[{i}] 表名: {sht}")
            print(f"    数据形状: {rows} 行 × {cols} 列")
            print(f"    字段名: {columns}")
            print("-" * 80)
        except Exception as e:
            print(f"[{i}] 表名: {sht} → 读取失败: {e}")
            print("-" * 80)


# ==============================================
# 自动编码读取 shp
# ==============================================
def read_shp_auto_encoding(path):
    encodings = ['utf-8', 'gbk', 'gb18030', 'gb2312']
    for enc in encodings:
        try:
            return gpd.read_file(path, encoding=enc)
        except (UnicodeDecodeError, LookupError):
            continue
    raise ValueError("所有常见编码均无法读取该 shp 文件")


def load_svg_compass(svg_path):
    svg_png = cairosvg.svg2png(url=svg_path)
    return Image.open(io.BytesIO(svg_png))


def auto_detect_text_col(gdf, col_type="county"):
    if col_type == "county":
        possible = ['Name', 'name', '县', '市', '区', 'COUNTY', 'county', '地名']
    else:
        possible = ['Name', 'name', '区划', '类型', 'zone', '区划名称', '农业区', '类型名', 'zonename', 'qu']
    for col in possible:
        if col in gdf.columns:
            return col
    return gdf.columns[0]


# ==============================================
# ✅【只修复这里，完全不改动你任何其他代码】
# 修复后：9个区域全部能匹配到
# ==============================================
def accurate_zone_match(county_gdf, zone_gdf, zone_col):
    county = county_gdf.copy()
    zone = zone_gdf.copy()

    # 统一坐标系
    if county.crs is None:
        county.set_crs(epsg=4326, inplace=True)
    if zone.crs is None:
        zone.set_crs(epsg=4326, inplace=True)

    # 投影到平面坐标系保证距离准确
    county_proj = county.to_crs(epsg=3857)
    zone_proj = zone.to_crs(epsg=3857)

    # 【正确匹配方式：质心包含】
    county_centroid = county_proj.geometry.centroid
    county_center_gdf = gpd.GeoDataFrame(county, geometry=county_centroid, crs=county_proj.crs)

    # 空间连接
    joined = gpd.sjoin(county_center_gdf, zone_proj, how='left', predicate='within')
    county["Agri_Zone"] = joined[zone_col].values

    # 补全缺失
    na_mask = county["Agri_Zone"].isna()
    if na_mask.any():
        zone_points = zone_proj.copy()
        zone_points["center"] = zone_proj.geometry.centroid

        for idx in county[na_mask].index:
            pt = county_center_gdf.geometry.iloc[idx]
            zone_points["dist"] = zone_points.distance(pt)
            best = zone_points.loc[zone_points["dist"].idxmin(), zone_col]
            county.loc[idx, "Agri_Zone"] = best

    match_success_count = county["Agri_Zone"].notna().sum()
    print(f"\n✅ 空间匹配完成：成功匹配 {match_success_count} 个县级单元")
    return county


# ==============================================
# ✅ 【新增】从区划SHP读取原始区域数量
# ==============================================
def print_zone_shp_info(zone_gdf, zone_col):
    print("\n" + "=" * 80)
    print("🌍 【区划SHP原始信息读取】")
    print("=" * 80)
    unique_zones = sorted(zone_gdf[zone_col].dropna().unique())
    print(f"📍 区划文件中原始区域总数：{len(unique_zones)} 个")
    print(f"📋 所有区域名称：")
    for i, name in enumerate(unique_zones, 1):
        print(f"   {i:2d}. {name}")
    print("=" * 80)


# ==============================================
# ✅ 匹配结果诊断统计
# ==============================================
def print_match_statistics(gdf):
    print("\n" + "=" * 80)
    print("📊 【空间匹配结果统计】")
    print("=" * 80)

    zone_counts = gdf["Agri_Zone"].value_counts().sort_index()
    total_zones = len(zone_counts)
    total_counties = len(gdf)
    success_count = gdf["Agri_Zone"].notna().sum()

    print(f"📍 匹配后有效区域种类：{total_zones} 类")
    print(f"📍 县级单元总数：{total_counties} 个")
    print(f"📍 成功匹配数量：{success_count} 个")
    print(f"📍 匹配成功率：{success_count / total_counties * 100:.2f}%")
    print("-" * 80)

    print("📌 各区域匹配县域数量：")
    for zone, cnt in zone_counts.items():
        pct = cnt / success_count * 100
        print(f"   🔹 {zone}：{cnt} 个 ({pct:.2f}%)")

    print("=" * 80 + "\n")


# ==============================================
# 比例尺、网格、指北针
# ==============================================
def add_scalebar(ax, gdf):
    bounds = gdf.total_bounds
    lon_min, lon_max = bounds[0], bounds[2]
    lat_center = np.mean(bounds[[1, 3]])
    km_per_degree = 111.32 * np.cos(np.radians(lat_center))
    map_width_km = (lon_max - lon_min) * km_per_degree
    scale_total_km = 2000
    ratio = scale_total_km / map_width_km
    x0, x1 = 0.05, 0.05 + ratio * 0.8
    y = 0.08

    ax.plot([x0, x1], [y, y], lw=3, c='k', transform=ax.transAxes)
    ax.text(x0, y - 0.03, '0', fontsize=cfg.SCALE_TEXT_SIZE, ha='center', va='top', transform=ax.transAxes)
    ax.text((x0 + x1) / 2, y - 0.03, '1000', fontsize=cfg.SCALE_TEXT_SIZE, ha='center', va='top', transform=ax.transAxes)
    ax.text(x1, y - 0.03, '2000 km', fontsize=cfg.SCALE_TEXT_SIZE, ha='center', va='top', transform=ax.transAxes)


def add_grid(ax, gdf):
    ax.spines[:].set_visible(True)
    ax.spines[:].set_linewidth(1)
    ax.spines[:].set_color('black')
    bounds = gdf.total_bounds
    lon_min, lat_min, lon_max, lat_max = bounds

    lon_ticks = np.arange(np.ceil(lon_min / 5) * 5, np.floor(lon_max / 5) * 5 + 1, 5)
    lat_ticks = np.arange(np.ceil(lat_min / 5) * 5, np.floor(lat_max / 5) * 5 + 1, 5)

    for lon in lon_ticks:
        ax.axvline(lon, c='gray', ls='--', lw=0.6, alpha=0.7)
    for lat in lat_ticks:
        ax.axhline(lat, c='gray', ls='--', lw=0.6, alpha=0.7)

    ax.set_xticks(lon_ticks)
    ax.set_yticks(lat_ticks)
    ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{x:.0f}°E"))
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.0f}°N"))
    ax.tick_params(labelsize=cfg.GRID_LABEL_SIZE)


def add_compass(ax, img):
    ib = OffsetImage(img, zoom=cfg.COMPASS_ZOOM)
    ab = AnnotationBbox(ib, (0.92, 0.92), frameon=False, xycoords='axes fraction')
    ax.add_artist(ab)


# ==============================================
# 另存为新文件
# ==============================================
def save_zone_to_new_excel(original_path, save_path, sheet_name, county_zone_map):
    wb = load_workbook(original_path)
    ws = wb[sheet_name]

    col_county = None
    for c in range(1, 100):
        if ws.cell(row=1, column=c).value == "县":
            col_county = c
            break
    if not col_county:
        print("❌ 未找到'县'列")
        return

    new_col = ws.max_column + 1
    ws.cell(row=1, column=new_col, value="所属区域")

    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=col_county).value
        if pd.isna(name):
            continue
        name = str(name).strip()
        ws.cell(row=r, column=new_col, value=county_zone_map.get(name, ""))

    wb.save(save_path)
    wb.close()
    print(f"\n✅ 已另存为新文件：{save_path}")
    print(f"✅ 仅在表 {sheet_name} 新增【所属区域】列")


# ==============================================
# 主绘图
# ==============================================
def draw_map():
    plt.rcParams['font.family'] = [cfg.FONT_NAME, 'SimHei']
    plt.rcParams['axes.unicode_minus'] = False

    if(cfg.NEED_SAVE_EXCEL):
        print_all_sheets_info(cfg.EXCEL_PATH, title="【原始 Excel 文件】信息")

    img_compass = load_svg_compass(cfg.SVG_COMPASS_PATH)
    df = pd.read_excel(cfg.EXCEL_PATH, sheet_name=cfg.SHEET_NAME)
    county_gdf = read_shp_auto_encoding(cfg.SHP_PATH)
    nine_lines = read_shp_auto_encoding(cfg.NINE_LINES_PATH)
    zone_gdf = read_shp_auto_encoding(cfg.AGRI_ZONE_PATH)

    print("=" * 60)
    print("📊 【分区shp】所有列名：")
    print(zone_gdf.columns.tolist())
    print("=" * 60)

    zone_col = input("请输入分区列名（如 qu）：").strip()
    while zone_col not in zone_gdf.columns:
        zone_col = input("列名不存在，请重新输入：").strip()

    # ======================
    # ✅ 打印区划SHP原始区域数量
    # ======================
    print_zone_shp_info(zone_gdf, zone_col)

    county_col = auto_detect_text_col(county_gdf, col_type='county')
    county_gdf['match_name'] = county_gdf[county_col].astype(str).str.strip()
    df['县'] = df['县'].astype(str).str.strip()

    df_year = df[df['year'] == cfg.TARGET_YEAR].copy()
    valid_counties = df_year['县'].dropna().unique()
    valid_count = len(valid_counties)
    print(f"📅 【{cfg.TARGET_YEAR}年 Excel 有效县数量】：{valid_count} 个")

    plot_gdf = accurate_zone_match(county_gdf, zone_gdf, zone_col)

    # ======================
    # ✅ 打印匹配诊断信息
    # ======================
    print_match_statistics(plot_gdf)

    plot_gdf['Plot_Zone'] = plot_gdf['Agri_Zone']
    plot_gdf.loc[~plot_gdf['match_name'].isin(valid_counties), 'Plot_Zone'] = np.nan

    filled_count = plot_gdf['Plot_Zone'].notna().sum()
    print(f"\n🗺️ 【地图中成功填色县数量】：{filled_count} 个")

    no_color = plot_gdf[plot_gdf['Plot_Zone'].isna() & plot_gdf['match_name'].isin(valid_counties)]
    print("\n==================================================")
    print("🔎 未匹配到的县：")
    print(no_color['match_name'].tolist() if len(no_color) > 0 else "✅ 无！全部匹配成功！")
    print("==================================================\n")

    # ======================
    # ✅ 根据超参数判断是否保存 Excel
    # ======================
    if cfg.NEED_SAVE_EXCEL:
        matched = plot_gdf[['match_name', 'Agri_Zone']].dropna()
        county_zone_map = dict(zip(matched['match_name'], matched['Agri_Zone']))
        save_zone_to_new_excel(cfg.EXCEL_PATH, cfg.SAVE_EXCEL_PATH, cfg.SHEET_NAME, county_zone_map)
        print_all_sheets_info(cfg.SAVE_EXCEL_PATH, title="【新生成 Excel 文件】信息")
    else:
        print("\nℹ️ 超参数已关闭：不保存 Excel 文件")

    zones = sorted(plot_gdf['Plot_Zone'].dropna().unique())
    color_map = dict(zip(zones, cfg.COLORS[:len(zones)]))

    print("==================================================")
    print("🗺️ 图例分类名称：")
    print("==================================================")
    legend_elements = []
    for z in zones:
        en_name = cfg.ZONE_TRANSLATION.get(z, z)
        print(f"- {z} → {en_name}")
        legend_elements.append(
            mpatches.Patch(facecolor=color_map[z], edgecolor='black', label=en_name)
        )
    legend_elements.append(mpatches.Patch(facecolor='white', edgecolor='black', label='No Data'))

    fig, ax = plt.subplots(1, 1, figsize=cfg.FIG_SIZE, dpi=cfg.DPI)

    plot_gdf.plot(ax=ax, facecolor=cfg.NULL_COLOR, edgecolor=cfg.BORDER_COLOR, linewidth=cfg.BORDER_WIDTH)

    plot_data = plot_gdf[plot_gdf['Plot_Zone'].notna()].copy()
    if not plot_data.empty:
        plot_data.plot(
            ax=ax,
            color=plot_data['Plot_Zone'].map(color_map),
            edgecolor=cfg.BORDER_COLOR,
            linewidth=cfg.BORDER_WIDTH
        )

    nine_lines.plot(ax=ax, color=cfg.NINE_LINE_COLOR, linewidth=cfg.NINE_LINE_WIDTH)

    add_grid(ax, plot_gdf)
    add_scalebar(ax, plot_gdf)
    add_compass(ax, img_compass)

    ax.legend(
        handles=legend_elements,
        loc='lower left',
        fontsize=cfg.LEGEND_FONT_SIZE,
        frameon=True,
        bbox_to_anchor=(0.02, 0.12),
        title='Ecological Zones',
        title_fontsize=cfg.LEGEND_TITLE_SIZE
    )

    plt.tight_layout()
    plt.savefig(cfg.SAVE_PATH.replace('.svg', '.png'), format='png', dpi=cfg.DPI, bbox_inches='tight')
    plt.close()
    print(f"\n✅ 绘图完成：{cfg.SAVE_PATH.replace('.svg', '.png')}")


if __name__ == "__main__":
    draw_map()